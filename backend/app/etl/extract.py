"""ETL Extract: read a dataset directory into RawQuestion dataclasses.

Pure file I/O + parsing. No DB, no business logic.

Two source shapes are supported:

1. **JSONL dataset** (PRD §10.3): a directory with ``manifest.json`` +
   ``questions.jsonl`` (the seeded osg10 shape). ``DatasetReader`` auto-detects
   this and reads it via the original manifest-driven path.
2. **Uploaded template** (PRD §10.1, FR-IMP-01): a directory with a single
   ``questions.csv`` / ``questions.xlsx`` / ``questions.json`` file produced by
   the ``POST /api/etl/upload`` endpoint. ``DatasetReader`` auto-detects the
   format from the file present and dispatches to ``CsvExtractor`` /
   ``XlsxExtractor`` / ``JsonExtractor``, which map the flat §10.1 column schema
   onto ``RawQuestion`` (carrying ``domain``/``knowledge_points``/``tags``/
   ``book`` in ``RawQuestion.meta`` for the transform/load stages).
"""

import csv
import hashlib
import json
import re
from dataclasses import dataclass, field
from pathlib import Path


@dataclass
class Bilingual:
    en: str
    zh: str


@dataclass
class RawSource:
    book: str
    edition: int
    section: str
    chapter: int
    chapter_title: str
    number: int


@dataclass
class RawOption:
    key: str
    text: Bilingual


@dataclass
class RawPromptItem:
    key: str
    text: Bilingual


@dataclass
class RawQuestion:
    id: str
    source: RawSource
    type: str
    stem: Bilingual
    options: list[RawOption]
    correct_keys: list[str]
    explanation: Bilingual
    meta: dict
    prompt_items: list[RawPromptItem] | None = None
    # Enrichment fields (PRD §10 import template / FR-ETL-09). All optional:
    # absence -> None, defaults applied downstream. ETL hardcoding these (#16/#18)
    # is what made CAT ability-matching meaningless and dropped per-option
    # explanations even when the source carried them.
    difficulty: int | None = None
    option_explanations: dict[str, Bilingual] | None = None
    license_status: str | None = None


@dataclass
class ExtractError:
    line_no: int | None
    external_id: str | None
    reason: str


def _bilingual(d: dict) -> Bilingual:
    return Bilingual(en=d.get("en", ""), zh=d.get("zh", ""))


# difficulty label -> int (PRD §11.1 range 1-5). Labels are a convenience for
# the CSV/XLSX/JSON import template (#35); int/numeric-string also accepted.
_DIFFICULTY_LABELS = {
    "very_easy": 1, "easy": 2, "medium": 3, "hard": 4, "very_hard": 5,
}


def _parse_difficulty(value) -> int | None:
    """Parse a source difficulty value to a clamped int in [1, 5], else None.

    Accepts int, numeric string, or label (easy/medium/hard/...). Out-of-range
    values clamp to 1 or 5. Garbage -> None (so the transform fallback applies).
    """
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip().lower()
        if v in _DIFFICULTY_LABELS:
            return _DIFFICULTY_LABELS[v]
        if v == "":
            return None
        try:
            n = int(v)
        except ValueError:
            return None
    elif isinstance(value, bool):  # bool is an int subclass; reject it
        return None
    elif isinstance(value, int):
        n = value
    else:
        return None
    return max(1, min(5, n))


def _parse_option_explanations(rec: dict) -> dict[str, Bilingual] | None:
    """Merge per-option explanations from the PRD §10 template fields.

    Supports two shapes:
      - split: ``option_explanations`` (en, {key: text}) + ``option_explanations_zh``
        (zh, {key: text}) - the documented CSV/XLSX/JSON template format.
      - nested: ``option_explanations`` as {key: {en, zh}}.
    Returns None when neither field is present. Keys with no text on either
    side still yield a Bilingual("", "") - transform treats empty as absent.
    """
    en_map = rec.get("option_explanations")
    zh_map = rec.get("option_explanations_zh")
    if en_map is None and zh_map is None:
        return None
    en_map = en_map or {}
    zh_map = zh_map or {}
    keys = set(en_map.keys()) | set(zh_map.keys())
    out: dict[str, Bilingual] = {}
    for k in keys:
        en_val = en_map.get(k, "")
        zh_val = zh_map.get(k, "")
        # A nested {en, zh} on either side is the alternate single-field shape.
        en_nested = en_val if isinstance(en_val, dict) else None
        zh_nested = zh_val if isinstance(zh_val, dict) else None
        en_str = (en_nested.get("en", "") if en_nested else (str(en_val) if en_val else ""))
        # zh: explicit zh_map wins, else pull from a nested en value, else a nested zh.
        if zh_nested:
            zh_str = zh_nested.get("zh", "")
        elif zh_val and not isinstance(zh_val, dict):
            zh_str = str(zh_val)
        elif en_nested:
            zh_str = en_nested.get("zh", "")
        else:
            zh_str = ""
        out[k] = Bilingual(en=en_str, zh=zh_str)
    return out


def _parse_record(rec: dict) -> RawQuestion:
    src = rec["source"]
    raw = RawQuestion(
        id=rec["id"],
        source=RawSource(
            book=src["book"],
            edition=src["edition"],
            section=src["section"],
            chapter=src["chapter"],
            chapter_title=src["chapter_title"],
            number=src["number"],
        ),
        type=rec["type"],
        stem=_bilingual(rec["stem"]),
        options=[
            RawOption(key=o["key"], text=_bilingual(o["text"]))
            for o in rec["options"]
        ],
        correct_keys=list(rec["correct_keys"]),
        explanation=_bilingual(rec["explanation"]),
        meta=rec.get("meta", {}),
        prompt_items=(
            [
                RawPromptItem(key=p["key"], text=_bilingual(p["text"]))
                for p in rec["prompt_items"]
            ]
            if rec.get("prompt_items")
            else None
        ),
        difficulty=_parse_difficulty(rec.get("difficulty")) or _parse_difficulty(rec.get("meta", {}).get("difficulty")),
        option_explanations=_parse_option_explanations(rec),
        license_status=(
            rec.get("license_status")
            or (rec.get("meta", {}) or {}).get("license_status")
        ),
    )
    return raw


class DatasetReader:
    """Read a dataset directory into RawQuestion records.

    Format is auto-detected from the directory contents (NOT the EtlDataset.format
    DB field, which is informational and historically ``json`` for JSONL datasets):

      - ``manifest.json`` + ``questions.jsonl``  -> JSONL dataset (PRD §10.3)
      - ``questions.csv``                        -> CSV template (PRD §10.1)
      - ``questions.xlsx``                       -> XLSX template (PRD §10.1)
      - ``questions.json``                       -> JSON template (a JSON array
                                                    of §10.1 row objects)

    Returns ``(raws, errors, content_hash)``. The content_hash covers whichever
    files are read, so drift detection between preview and commit works for both
    shapes.
    """

    def __init__(self, dataset_path: str | Path):
        self.path = Path(dataset_path)

    def read(self) -> tuple[list[RawQuestion], list[ExtractError], str]:
        if (self.path / "manifest.json").exists() and (self.path / "questions.jsonl").exists():
            return self._read_jsonl()
        for filename, extractor_cls in (
            ("questions.csv", CsvExtractor),
            ("questions.xlsx", XlsxExtractor),
            ("questions.json", JsonExtractor),
        ):
            if (self.path / filename).exists():
                raws, errors = extractor_cls(self.path / filename).read()
                return raws, errors, self._content_hash([filename])
        raise FileNotFoundError(
            f"no recognized dataset file in {self.path} "
            f"(expected manifest.json+questions.jsonl, questions.csv, .xlsx, or .json)"
        )

    def _read_jsonl(self) -> tuple[list[RawQuestion], list[ExtractError], str]:
        raws: list[RawQuestion] = []
        errors: list[ExtractError] = []
        content_hash = self._content_hash(["manifest.json", "questions.jsonl"])

        manifest = json.loads((self.path / "manifest.json").read_text())
        expected = manifest.get("total_questions")

        jsonl = self.path / "questions.jsonl"
        with jsonl.open() as f:
            for line_no, line in enumerate(f, start=1):
                line = line.strip()
                if not line:
                    continue
                rec = None
                try:
                    rec = json.loads(line)
                    raws.append(_parse_record(rec))
                except (json.JSONDecodeError, KeyError, TypeError) as exc:
                    external_id = rec.get("id") if isinstance(rec, dict) else None
                    errors.append(
                        ExtractError(
                            line_no=line_no,
                            external_id=external_id,
                            reason=f"{type(exc).__name__}: {exc}",
                        )
                    )

        if expected is not None and expected != len(raws):
            errors.append(
                ExtractError(
                    line_no=None,
                    external_id=None,
                    reason=f"manifest total_questions={expected} but parsed {len(raws)} records",
                )
            )

        return raws, errors, content_hash

    def _content_hash(self, names: list[str]) -> str:
        h = hashlib.sha256()
        for name in names:
            h.update(name.encode())
            h.update((self.path / name).read_bytes())
        return h.hexdigest()


# ---------------------------------------------------------------------------
# Uploaded-template extractors (PRD §10.1, FR-IMP-01 / #35).
#
# CSV/XLSX/JSON uploads all share the flat §10.1 column schema, so they funnel
# through one ``_row_to_raw`` mapper. Each extractor only differs in how it
# turns the file into a stream of row dicts.
# ---------------------------------------------------------------------------

_OPTION_LETTERS = ("a", "b", "c", "d", "e", "f")


def _parse_int(value, default: int | None = None) -> int | None:
    if value is None or value == "":
        return default
    try:
        return int(str(value).strip())
    except (ValueError, TypeError):
        return default


def _parse_chapter_number(value) -> int:
    """Extract a chapter number from a free-text cell like 'Chapter 1' / '第3章'.

    Returns 0 when no digit is found (the chapter is informational only)."""
    if not value:
        return 0
    m = re.search(r"\d+", str(value))
    return int(m.group()) if m else 0


def _split_list(value) -> list[str]:
    """Split a ';' or ',' separated cell into trimmed non-empty items."""
    if not value:
        return []
    return [s.strip() for s in re.split(r"[;,]", str(value)) if s.strip()]


def _normalize_row(row: dict) -> dict:
    """Parse JSON-string cells (option_explanations[_zh]) into dicts.

    CSV/XLSX cells arrive as strings; JSON uploads arrive already-typed. An
    invalid JSON string is dropped (None) so the row still imports without
    per-option explanations rather than failing the whole batch.
    """
    for key in ("option_explanations", "option_explanations_zh"):
        v = row.get(key)
        if isinstance(v, str) and v.strip():
            try:
                row[key] = json.loads(v)
            except json.JSONDecodeError:
                row[key] = None
    return row


def _cell(row: dict, key: str) -> str:
    """Case-insensitive, whitespace-trimmed string cell lookup."""
    if key in row:
        val = row[key]
    else:
        # tolerant header lookup: match case-insensitively
        val = next((v for k, v in row.items() if isinstance(k, str) and k.lower() == key), "")
    return (str(val).strip() if val is not None else "")


def _row_to_raw(row: dict, row_no: int) -> RawQuestion:
    """Map one PRD §10.1 template row onto a RawQuestion.

    Carries the per-question ``domain`` / ``knowledge_points`` / ``tags`` /
    ``book`` / ``edition`` columns in ``meta`` (the JSONL shape has no such
    fields - domain is derived via ChapterDomainMapping - so the transform/load
    stages treat absent meta values as 'use the JSONL behavior').
    """
    stem_en = _cell(row, "question_text")
    stem_zh = _cell(row, "question_text_zh")
    qtype = _cell(row, "question_type") or "single_choice"

    options: list[RawOption] = []
    for letter in _OPTION_LETTERS:
        en = _cell(row, f"option_{letter}")
        if not en:
            break  # options are contiguous a,b,(c),(d); stop at the first gap
        zh = _cell(row, f"option_{letter}_zh")
        options.append(RawOption(key=letter.upper(), text=Bilingual(en=en, zh=zh)))

    correct = [c.upper() for c in _split_list(_cell(row, "correct_answers"))]

    # Content-addressed external id: stable across re-uploads of the same file
    # (so a re-import is idempotent - 'unchanged' or 'updated', not re-created)
    # and independent of row order.
    ext_id = "upload-" + hashlib.sha1(f"{stem_en}|{qtype}".encode("utf-8")).hexdigest()[:16]

    book = _cell(row, "book") or "User Import"
    edition = _parse_int(_cell(row, "edition"), default=1) or 1
    chapter_raw = _cell(row, "chapter")
    chapter_num = _parse_chapter_number(chapter_raw)
    domain = _parse_int(_cell(row, "domain"))
    knowledge_points = _split_list(_cell(row, "knowledge_points"))
    tags = _split_list(_cell(row, "tags"))

    meta = {
        "domain": domain,
        "knowledge_points": knowledge_points,
        "tags": tags,
        "book": book,
        "edition": edition,
        "source_label": _cell(row, "source") or "user_import",
    }

    return RawQuestion(
        id=ext_id,
        source=RawSource(
            book=book,
            edition=edition,
            section="",
            chapter=chapter_num,
            chapter_title=chapter_raw,
            number=row_no,
        ),
        type=qtype,
        stem=Bilingual(en=stem_en, zh=stem_zh),
        options=options,
        correct_keys=correct,
        explanation=Bilingual(
            en=_cell(row, "explanation"),
            zh=_cell(row, "explanation_zh"),
        ),
        meta=meta,
        prompt_items=None,
        difficulty=_parse_difficulty(_cell(row, "difficulty")),
        option_explanations=_parse_option_explanations(_normalize_row(row)),
        license_status=(_cell(row, "license_status") or None),
    )


class _TemplateExtractor:
    """Shared read loop: parse file -> row dicts -> RawQuestion records.

    Subclasses implement ``_iter_rows()`` yielding (row_no, row_dict)."""

    def __init__(self, path: str | Path):
        self.path = Path(path)

    def read(self) -> tuple[list[RawQuestion], list[ExtractError]]:
        raws: list[RawQuestion] = []
        errors: list[ExtractError] = []
        try:
            iterable = self._iter_rows()
            for row_no, row in iterable:
                try:
                    raws.append(_row_to_raw(row, row_no))
                except Exception as exc:
                    errors.append(
                        ExtractError(
                            line_no=row_no,
                            external_id=None,
                            reason=f"{type(exc).__name__}: {exc}",
                        )
                    )
        except Exception as exc:
            # File-level/structural failure (e.g. JSON isn't an array, XLSX
            # corrupt). Surface as one error rather than crashing the batch.
            errors.append(
                ExtractError(
                    line_no=None,
                    external_id=None,
                    reason=f"{type(exc).__name__}: {exc}",
                )
            )
        return raws, errors


class CsvExtractor(_TemplateExtractor):
    def _iter_rows(self):
        with self.path.open(newline="", encoding="utf-8-sig") as f:  # utf-8-sig strips BOM
            # enumerate from 2: row 1 is the header in a CSV
            for i, row in enumerate(csv.DictReader(f), start=2):
                yield i, {k: ("" if v is None else v) for k, v in row.items() if k is not None}


class JsonExtractor(_TemplateExtractor):
    def _iter_rows(self):
        data = json.loads(self.path.read_text(encoding="utf-8-sig"))
        if isinstance(data, dict):
            data = data.get("questions") or data.get("data")
            if data is None:
                raise ValueError(
                    "JSON upload must be an array of question rows (or {questions: [...]})"
                )
        if not isinstance(data, list):
            raise ValueError(
                "JSON upload must be an array of question rows (or {questions: [...]})"
            )
        for i, row in enumerate(data, start=1):
            if isinstance(row, dict):
                yield i, dict(row)


class XlsxExtractor(_TemplateExtractor):
    def _iter_rows(self):
        # Imported lazily so the openpyxl dep is only required for XLSX uploads
        # (CSV/JSON use the stdlib and work without it).
        from openpyxl import load_workbook

        wb = load_workbook(self.path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration:
                return
            headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
            for i, values in enumerate(rows, start=2):
                if values is None or all(v is None for v in values):
                    continue
                yield i, {
                    headers[j]: ("" if j >= len(values) or values[j] is None else values[j])
                    for j in range(len(headers))
                    if headers[j]
                }
        finally:
            wb.close()
